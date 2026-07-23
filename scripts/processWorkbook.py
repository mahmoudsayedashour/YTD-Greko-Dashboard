"""
Greko Egypt — Workbook Preprocessor (Python/openpyxl version)
Produces the exact same processed-data.json v3 format as processWorkbook.js.

Usage:
    python scripts/processWorkbook.py

openpyxl with read_only=True is 10-20x faster than the Node xlsx library
for large sheets because it streams rows without loading all of them
into memory at once.
"""

import json
import math
import os
import re
import sys
import urllib.request
from datetime import datetime

# Fix Windows console encoding (cp1252 → utf-8)
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

# ── Config ────────────────────────────────────────────────────────────────────
WORKBOOK_URL = (
    "https://kpvezuvifxoatyen.public.blob.vercel-storage.com/"
    "New%20Microsoft%20Excel%20Worksheet.xlsx"
)
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "processed-data.json")
TMP_PATH    = os.path.join(os.path.dirname(__file__), "..", "data", "_tmp_workbook.xlsx")

MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MONTHS_FULL  = ["January","February","March","April","May","June",
                "July","August","September","October","November","December"]
MONTHS_AR = [
    "يناير","فبراير","مارس","أبريل","مايو","يونيو",
    "يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر",
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def sf(v):
    try: return float(v) if v is not None else 0.0
    except (TypeError, ValueError): return 0.0

def ss(v):
    return str(v).strip() if v is not None else ""

def r4(v): return round(v, 4)
def ri(v): return int(round(v)) if v is not None else 0

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v
    if isinstance(v, str):
        v = v.strip()
        if not v: return None
        # Try to parse 'Monday, December 29, 2025'
        # Actually, simpler: regex for Month YYYY
        try:
            # Fallback if standard format
            return datetime.strptime(v, "%A, %B %d, %Y")
        except ValueError:
            pass
        
        # Try finding a month name
        for i, m in enumerate(MONTHS_FULL):
            if m.lower() in v.lower():
                # find year
                yr_match = re.search(r'\b(20\d\d)\b', v)
                yr = int(yr_match.group(1)) if yr_match else 0
                return datetime(yr, i+1, 1)
    return None

def row_month(v):
    d = parse_date(v)
    return d.month if d else 0

def row_year(v):
    d = parse_date(v)
    return d.year if d else 0

# ── String intern table ───────────────────────────────────────────────────────
class StringTable:
    def __init__(self):
        self._map = {}
        self._arr = []
    def intern(self, s):
        s = ss(s)
        if s in self._map: return self._map[s]
        idx = len(self._arr)
        self._arr.append(s)
        self._map[s] = idx
        return idx
    def to_list(self): return self._arr

# ── Download ──────────────────────────────────────────────────────────────────
def download(url, dest):
    print(f"📥 Downloading workbook…")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req) as r, open(dest, "wb") as f:
        total = int(r.headers.get("Content-Length", 0))
        downloaded = 0
        while True:
            chunk = r.read(65536)
            if not chunk: break
            f.write(chunk)
            downloaded += len(chunk)
            if total:
                sys.stdout.write(f"\r   {downloaded/1024/1024:.1f} / {total/1024/1024:.1f} MB")
                sys.stdout.flush()
    size = os.path.getsize(dest)
    print(f"\n   Total: {size/1024/1024:.2f} MB")

# ── Forecast parser ───────────────────────────────────────────────────────────
def parse_forecast(ws):
    """Parse a Forecast sheet → {code: {1..12: {ton, carton, cups}}}"""
    result = {}
    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [ss(c) for c in row]
            continue
        if not any(c is not None for c in row): continue
        row_dict = {h: v for h, v in zip(headers, row)}
        code = ss(row_dict.get("Code") or row_dict.get("code") or "")
        if not code: continue
        result[code] = {}
        for m in range(1, 13):
            en = MONTHS_FULL[m-1]
            ar = MONTHS_AR[m-1]
            ton = sf(
                row_dict.get(ar + " طن") or row_dict.get("طن " + ar) or
                row_dict.get(en + " Ton") or row_dict.get(en + "Ton") or 0
            )
            carton = sf(
                row_dict.get(ar + " كراتين") or
                row_dict.get(en + " Cartons") or row_dict.get(en + "Cartons") or
                row_dict.get(en + " Carton") or 0
            )
            cups = sf(row_dict.get(en + " Cups") or row_dict.get(en + "Cups") or 0)
            result[code][m] = {"ton": ton, "carton": carton, "cups": cups}
    print(f"   ✓ forecast: {len(result)} product codes")
    return result

# ── Actual parser ─────────────────────────────────────────────────────────────
def parse_actual(ws, channel_map, ST, class_map, manager_map, outlet_map, label):
    """Stream-parse an Actual sheet → list of compact rows"""
    out = []
    skipped_nodate = 0
    skipped_zero   = 0

    headers = None
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [ss(c) for c in row]
            print(f"   [DEBUG] ALL Headers for {label}: {headers}")
            continue
        if not any(c is not None for c in row): continue

        rd = dict(zip(headers, row))
        # (Debug prints removed)

        month = row_month(rd.get("Delivery Date"))
        if not month: skipped_nodate += 1; continue

        tn_raw = sf(rd.get("Num Ton"))
        ct_raw = sf(rd.get("Num Carton"))
        cp_raw = sf(rd.get("Invoice lines/Quantity"))
        if tn_raw == 0 and ct_raw == 0 and cp_raw == 0:
            skipped_zero += 1; continue

        tn = r4(tn_raw)
        ct = r4(ct_raw)
        cp = ri(cp_raw)

        partner = ss(rd.get("Invoice Partner Display Name.1") or
                     rd.get("Invoice lines/Partner") or
                     rd.get("Invoice Partner Display Name") or
                     rd.get("Partner") or "")
        tag = ss(rd.get("Tags") or rd.get("tags") or
                 rd.get("Classification") or rd.get("Customer Category") or "")
        ch  = ss(rd.get("Channel") or rd.get("channel") or
                 rd.get("Trade Channel") or rd.get("Sales Channel") or "")

        if partner and tag: class_map[partner] = tag
        if partner and ch:
            channel_map[partner] = ch
            norm = " ".join(partner.split())
            if norm != partner: channel_map[norm] = ch
            m2 = re.match(r"^\[([^\]]+)\]", partner)
            if m2: channel_map["__code__" + m2.group(1)] = ch
            
        sm = ss(rd.get("Sales Manager") or "")
        pe = ss(rd.get("Partner English") or "")
        if partner and sm: manager_map[partner] = sm
        if partner and pe: outlet_map[partner] = pe

        code = ss(rd.get("Code") or "")
        inv_type = ss(rd.get("Invoice lines/Number Type") or "")
        ref      = ss(rd.get("Invoice lines/Reference") or "")
        rf_r = 1 if (ref and ref[0].upper() == "R") else 0

        out.append([
            month,
            ST.intern(code),
            ST.intern(partner),
            ST.intern(inv_type),
            rf_r,
            tn, ct, cp,
        ])

    print(f"   ✓ {label}: {len(out)} rows "
          f"({skipped_nodate} no-date, {skipped_zero} all-zero skipped)")
    return out

# ── DAX validation ────────────────────────────────────────────────────────────
def dax_validate(rows, strings, label):
    t_sum = t_partial = t_rinv = 0.0
    for r in rows:
        tn, rf_r, typ = r[5], r[4], strings[r[3]]
        t_sum += tn
        if rf_r: t_partial += abs(tn)
        if typ == "RINV": t_rinv += tn
    t_ret  = abs(t_rinv - t_partial)
    t_sale = t_sum - t_partial - t_ret
    print(f"\n  📊 DAX Preview — {label}")
    print(f"     Raw ΣTon      : {t_sum:.2f}")
    print(f"     Partial Returns: {t_partial:.2f}")
    print(f"     Returns (RINV) : {t_ret:.2f}")
    print(f"     → Sales Ton    : {t_sale:.2f}")

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("\n═══════════════════════════════════════════════════")
    print(" Greko Egypt – Workbook Preprocessor  (Python/openpyxl)")
    print("═══════════════════════════════════════════════════\n")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    # 1. Download
    download(WORKBOOK_URL, TMP_PATH)

    # 2. Open with openpyxl in read-only + data-only mode (fastest)
    print("\n📊 Opening workbook (read_only, data_only)…")
    wb = openpyxl.load_workbook(TMP_PATH, read_only=True, data_only=True)
    print(f"   Sheets: {wb.sheetnames}")

    # Resolve sheet names
    def find_sheet(pattern, default):
        for n in wb.sheetnames:
            if re.search(pattern, n, re.I): return n
        return default

    sheet_act25 = find_sheet(r"Actual.?25$",       "Actual 25")
    sheet_act26 = find_sheet(r"Actual.?2?0?26$",   "Actual 26")
    sheet_fct25 = find_sheet(r"Forecast.?25$",     "Forecast 25")
    sheet_fct26 = find_sheet(r"Forecast.?2?0?26$", "Forecast 26")

    required = ["Main Data", "Customers", sheet_fct25, sheet_fct26, sheet_act25, sheet_act26]
    missing  = [s for s in required if s not in wb.sheetnames]
    if missing:
        print(f"\n❌  Missing: {missing}  |  Available: {wb.sheetnames}")
        sys.exit(1)
    print(f"   Using: {sheet_act25} / {sheet_act26} / {sheet_fct25} / {sheet_fct26}")

    # 3. Build lookup maps from Main Data + Customers
    print("\n🗺  Building lookup maps…")
    product_map  = {}
    category_map = {}
    channel_map  = {}
    class_map    = {}
    manager_map  = {}
    outlet_map   = {}

    ws_main = wb["Main Data"]
    main_headers = None
    for row in ws_main.iter_rows(values_only=True):
        if main_headers is None:
            main_headers = [ss(c) for c in row]
            print(f"   [DEBUG] Main Data headers: {main_headers}")
            continue
        rd = dict(zip(main_headers, row))
        code = ss(rd.get("Code") or rd.get("code") or "")
        if not code: continue
        # Exact column names as per the workbook
        name = ss(
            rd.get("Invoice lines/Product") or
            rd.get("Product Name") or rd.get("ProductName") or
            rd.get("Name") or code
        )
        cat  = ss(
            rd.get("Product Category") or
            rd.get("Category") or rd.get("category") or ""
        )
        if name: product_map[code]  = name
        if cat:  category_map[code] = cat
    print(f"   Main Data: {len(product_map)} products, {len(set(category_map.values()))} categories")

    ws_cust = wb["Customers"]
    cust_headers = None
    for row in ws_cust.iter_rows(values_only=True):
        if cust_headers is None:
            cust_headers = [ss(c) for c in row]; continue
        rd = dict(zip(cust_headers, row))
        cust = ss(rd.get("Customer") or rd.get("Name") or rd.get("Partner") or "")
        ch   = ss(rd.get("Channel") or rd.get("channel") or "")
        tag  = ss(rd.get("Tags") or rd.get("Classification") or "")
        if cust and ch:
            channel_map[cust] = ch
            norm = " ".join(cust.split())
            if norm != cust: channel_map[norm] = ch
            m2 = re.match(r"^\[([^\]]+)\]", cust)
            if m2: channel_map["__code__" + m2.group(1)] = ch
        if cust and tag: class_map[cust] = tag
    print(f"   Customers: {len(channel_map)} channel mappings")

    # 4. Parse forecast sheets
    print("\n📈 Parsing forecast sheets…")
    fc25 = parse_forecast(wb[sheet_fct25])
    fc26 = parse_forecast(wb[sheet_fct26])

    # 5. Parse actual sheets (streaming — handles 700K rows efficiently)
    print("\n📦 Parsing actual sheets (streaming)…")
    ST = StringTable()
    rows25 = parse_actual(wb[sheet_act25], channel_map, ST, class_map, manager_map, outlet_map, sheet_act25)
    rows26 = parse_actual(wb[sheet_act26], channel_map, ST, class_map, manager_map, outlet_map, sheet_act26)
    wb.close()

    strings = ST.to_list()
    print(f"\n   String table: {len(strings)} entries")

    # 6. DAX validation
    dax_validate(rows25, strings, "2025")
    dax_validate(rows26, strings, "2026")

    # 7. Write output
    print(f"\n💾 Writing {OUTPUT_PATH}…")
    payload = {
        "version":     3,
        "generated":   datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source":      WORKBOOK_URL,
        "strings":     strings,
        "productMap":  product_map,
        "categoryMap": category_map,
        "channelMap":  channel_map,
        "classMap":    class_map,
        "managerMap":  manager_map,
        "outletMap":   outlet_map,
        "fc25":        fc25,
        "fc26":        fc26,
        "rows25":      rows25,
        "rows26":      rows26,
    }
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    size_mb = os.path.getsize(OUTPUT_PATH) / 1024 / 1024
    print(f"   ✅  Done! {size_mb:.1f} MB written.")
    print(f"   rows25={len(rows25)}  rows26={len(rows26)}  strings={len(strings)}")

    # 8. Clean up temp file
    try: os.remove(TMP_PATH)
    except: pass

if __name__ == "__main__":
    main()
